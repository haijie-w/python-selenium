#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2019/8/13 16:42
# @Author : whj

# 封装部分维护在此
import openpyxl
import xlrd
import time
from config.config_h5 import basic_config
from config.config_h5 import element_config
from common.logger import *
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from xlrd import xldate_as_tuple
from datetime import *
from time import sleep
from xlutils.copy import copy
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from selenium import webdriver


# 断言判断类
class Assert():
	logger = Logger()

	@classmethod
	def __init__(cls, driver):
		cls.driver = driver

	# 判断xpath是否存在，若存在返回True，否则返回False
	# Kawy @2017-08-25 Changed
	@classmethod
	def assertXPathExistByXPath(cls, page, element):
		try:
			WebHandle.byXPath(page, element)
			cls.logger.mylogger('Element is exist!')
			return True
		except Exception, e:
			cls.logger.mylogger('Element does not exist!')
			# raise Exception
			return False

	# 判断xpath是否存在，若存在返回True，否则返回False
	# Kawy @2017-11-08 Created
	@classmethod
	def assertXPathExist(cls, path):
		try:
			cls.driver.find_element_by_xpath(path)
			cls.logger.mylogger('Element is exist!')
			return True
		except Exception, e:
			cls.logger.mylogger('Element does not exist!')
			# raise Exception
			return False

	# 判断id是否存在，若存在返回True，否则返回False
	# Kawy @2017.08.23 Changed
	# 不存在return False,在调用该函数的地方raise Exception，因为要刷新页面逻辑难处理
	@classmethod
	def assertElementExistByID(cls, page, element):
		try:
			WebHandle.byId(page, element)
			cls.logger.mylogger(element + ' is exist!')
			return True
		except Exception, e:
			cls.logger.mylogger(element + ' does not exist!', -1)
			# raise Exception
			return False

	# 判断id是否不存在，若不存在返回True，否则返回False
	# Kawy @2017-10-25 Createed
	# 存在return False,在调用该函数的地方raise Exception，因为要刷新页面逻辑难处理
	@classmethod
	def assertElementNotExistByID(cls, page, element):
		try:
			WebHandle.byId(page, element)
			cls.logger.mylogger(element + ' is exist!', -1)
			return True
		except Exception, e:
			cls.logger.mylogger(element + ' does not exist!')
			# raise Exception
			return False

	# Zoe @2017-08-10
	# 判断页面上是否存在某个特定的文字，若存在返回True，否则返回False

	@classmethod
	def assertTextExist(cls, page, element, str):
		ele = WebHandle.getElementText(page, element)
		if str == ele:
			cls.logger.mylogger(str + 'is exist!')
			return True
		else:
			cls.logger.mylogger(str + ' does not exist!', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	# return False

	# Zoe @2017-08-10
	# 判断策略管理页面创建成功后是否有数据，若存在返回True，否则返回False

	@classmethod
	def assertStrategyExist(cls):
		# text=cls.driver.find_element_by_class_name(classname).text
		if str(cls.driver.find_element_by_class_name('list').text).split('：')[1].split('】')[0].split('【')[1] and (
				str(cls.driver.find_element_by_class_name('list').text).split('：')[1].split('】')[0].split('【')[1]).find(
			u'暂无数据') == -1:
			cls.logger.mylogger('策略ID=' +
								str(cls.driver.find_element_by_class_name('list').text).split('：')[1].split('】')[
									0].split('【')[1])
		else:
			cls.logger.mylogger('策略ID为空', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception
		if str(cls.driver.find_element_by_class_name('list').text).split('：')[2].split('获')[0].split('\n')[0] and (
				str(cls.driver.find_element_by_class_name('list').text).split('：')[2].split('获')[0].split('\n')[
					0]).find(u'暂无数据') == -1:
			cls.logger.mylogger(
				str('Token=' + cls.driver.find_element_by_class_name('list').text).split('：')[2].split('获')[0].split(
					'\n')[0])
		else:
			cls.logger.mylogger('Token为空', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception
		if str(cls.driver.find_element_by_class_name('list').text).split('：')[3].split('执')[0].split('\n')[0] and (
				str(cls.driver.find_element_by_class_name('list').text).split('：')[3].split('执')[0].split('\n')[
					0]).find(u'暂无数据') == -1:
			cls.logger.mylogger('数据接口=' +
								str(cls.driver.find_element_by_class_name('list').text).split('：')[3].split('执')[
									0].split('\n')[0])

		else:
			cls.logger.mylogger('数据接口为空', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception
		if str(cls.driver.find_element_by_class_name('list').text).split('：')[4].split('回')[0].split('\n')[0] and (
				str(cls.driver.find_element_by_class_name('list').text).split('：')[4].split('回')[0].split('\n')[
					0]).find(u'暂无数据') == -1:
			cls.logger.mylogger('交易接口=' +
								str(cls.driver.find_element_by_class_name('list').text).split('：')[4].split('回')[
									0].split('\n')[0])
		else:
			cls.logger.mylogger('交易接口为空', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception
		if str(cls.driver.find_element_by_class_name('list').text).split('：')[5].split('\n')[0] and (
				str(cls.driver.find_element_by_class_name('list').text).split('：')[5].split('\n')[0]).find(
			u'暂无数据') == -1:
			cls.logger.mylogger(
				'回测区间=' + str(cls.driver.find_element_by_class_name('list').text).split('：')[5].split('\n')[0])
		else:
			cls.logger.mylogger('回测区间为空', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	# 字符串信息比较，返回比较结果正确与否
	@classmethod
	def assertInfoEquals(cls, str1, str2):
		if str1 == str2:
			cls.logger.mylogger("Info is matching")
			return True
		else:
			cls.logger.mylogger("Info is mismatching", -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception  # return False

	# 判断表、图、ul是否存在
	@classmethod
	def assertTSUExistByXpath(cls, page, element, str_error):
		if not cls.assertXPathExistByXPath(page, element):
			try:
				cflag = False  # 判断有无图的标兵
				for i in range(2):
					if cls.assertXPathExistByXPath(page, element):
						cflag = True
						break
					else:
						sleep(1)
				assert cflag is True
				cls.logger.mylogger(str_error + 'OK')
				return True
			except Exception, e:
				cls.logger.mylogger(str_error + 'ERROR', -1)
				html = Logger.creatHtml(picpath, pics)
				Logger.mylogger(html)
				return False
		else:
			cls.logger.mylogger(str_error + 'OK')
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			return True

	# 第i个信息不匹配
	@classmethod
	def assertErrorType(cls, i):
		if i == 1:
			return " factor mismatching"
		elif i == 2:
			return " time mismatching"
		elif i == 3:
			return " stock pool mismatching"
		elif i == 4:
			return " benchmark mismatching"
		elif i == 5:
			return " group mismatching"
		elif i == 6:
			return " location frequency mismatching"
		elif i == 7:
			return " factor direction mismatching"
		cls.logger.mylogger('第' + str(i) + '条信息不匹配', -1)
		html = Logger.creatHtml(picpath, pics)
		Logger.mylogger(html)
		raise Exception

	# 判断id对应的弹出框是否存在，等待若干秒期间内，如果弹出框存在，返回True，否则返回False
	# Kawy @2017.08.23 Changed
	# 判断正在回测的提示框，不存在的情况下不抛出异常，改为在调用该函数的地方视情况判断是否抛出异常
	def assertPromptBox(cls, page, element):
		for i in range(50):
			if cls.assertElementExistByID(page, element):
				sleep(1)
			else:
				cls.logger.mylogger("PromptBox is not exist")
				# raise Exception
				return False
		cls.logger.mylogger("PromptBox is exist", -1)
		html = Logger.creatHtml(picpath, pics)
		Logger.mylogger(html)
		return True

	# Zoe @2017-08-10
	# 字符串信息比较不一致，不一致则返回true
	@classmethod
	def assertInfoNotEquals(cls, str1, str2):
		if str1 == str2:
			cls.logger.mylogger(str1 + " is equal to " + str2, -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception  # return False
		else:
			cls.logger.mylogger(str1 + " is not equal to " + str2)
			return True

	# Zoe @2017-08-22
	# 判断整个页面上是否存在某个特定的文字，若存在返回True，否则返回False

	@classmethod
	def assertPageTextExist(cls, path):
		try:
			ele = cls.driver.find_element_by_xpath(path)
			cls.logger.mylogger(path + 'is exist!')
			return True
		except Exception, e:
			cls.logger.mylogger(path + ' does not exist!', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	# Zoe @2017-08-22
	# 判断整个页面上是否存在某个特定的文字，若不存在返回True，否则返回False

	@classmethod
	def assertPageTextNotExist(cls, path):
		falg = True
		try:
			ele = cls.driver.find_element_by_xpath(path)
			falg = False
		except:
			cls.logger.mylogger(path + ' does not exist!')
			return True
		if not falg:
			cls.logger.mylogger(path + 'is exist!', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	# Zoe @2017-08-22
	# 判断整个页面上是否存在某个特定的文字，若不存在返回True，否则返回False

	'''Zoe @2017-09-3 Created
	根据xpath属性和table中的tr定位，并检查是否包含特定的元素
	text:需要检查的唯一的值
	page, element:xpath选取表格
	'''

	@classmethod
	def assertTableExsitByText(cls, text, page, element):
		flag = False
		table_tr_list = []
		# 选中table下所有的tr开头的元素赋值到列表中
		table_tr_list = WebHandle.byXPath(page, element).find_elements_by_tag_name("tr")
		# for选好获取tr元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 判断是否包含text
			if text in table_text:
				cls.logger.mylogger(text + ' is exist!')
				flag = True
				break
		if not flag:
			cls.logger.mylogger(text + 'does not exist!', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	'''Zoe @2017-09-3 Created
		根据xpath属性和table中的tr定位，并检查是否包含特定的元素
		text:需要检查的唯一的值
		 page, element:xpath选取表格
	   '''

	@classmethod
	def assertTableNotExsitByText(cls, text, page, element):
		flag = False
		table_tr_list = []
		# 选中table下所有的tr开头的元素赋值到列表中
		table_tr_list = WebHandle.byXPath(page, element).find_elements_by_tag_name("tr")
		# for选好获取tr元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 判断是否包含text
			if text in table_text:
				flag = True
				break
		if flag:
			cls.logger.mylogger(text + 'is exist!', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception
		else:
			cls.logger.mylogger(text + ' does not exist!')

	# Kawy @2017-09-01 Changed
	# ####判断图表中是否存在某个元素，queryContent表示要识别的元素的值，page-element表示到table的xpath####
	@classmethod
	def assertTableTextExist(cls, queryContent, page, element):
		a = '\n'
		flag = False
		# 按行查询表格的数据，取出的数据是一整行，按空格分隔每一列的数据
		table_tr_list = WebHandle.byXPath(page, element).text
		if a in table_tr_list:
			list = table_tr_list.split('\n')  # 以\n拆分成若干个(个数与列的个数相同)一维列表
			# 循环遍历table数据，确定查询数据的位置
			for i in range(len(list)):
				arr = list[i].split(" ")
				for j in range(len(arr)):
					if queryContent == arr[j]:
						Logger.mylogger(queryContent + '存在！')
						return True
		else:
			# 表格只有一行
			list = table_tr_list.split(" ")
			for i in range(len(list)):
				arr = list[i].split(" ")
				for j in range(len(arr)):
					if queryContent == arr[j]:
						Logger.mylogger(queryContent + '存在！')
						return True
		# 判读queryContent不存在，截图打印日志
		Logger.mylogger(queryContent + '不存在！', -1)
		return False

	'''Zoe @2017-09-3 Created
		根据xpath属性和list中的li定位，并检查是否包含特定的元素
		text:需要检查的唯一的值
		 page, element:xpath选取整个list
	   '''

	@classmethod
	def assertLiExsitByTest(cls, text, page, element):
		flag = False
		table_tr_list = []
		# 选中table下上所有的tr开头的元素赋值到列表中
		table_tr_list = cls.driver.find_element_by_xpath(local_config[page][element][1]).find_elements_by_tag_name("li")
		# for选好获取tr元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 判断是否包含text
			if text in table_text:
				cls.logger.mylogger(table_text + ' is fuzzy search with ' + text)
				flag = True
				continue
			else:
				cls.logger.mylogger(table_text + ' is  not fuzzy search with ' + text, -1)
				html = Logger.creatHtml(picpath, pics)
				Logger.mylogger(html)
				raise Exception


# 表格操作类
class ExcelHandle():
	logger = Logger()

	@classmethod
	def __init__(cls, driver):
		cls.driver = driver

	# 创建保存表格数据到..\Data目录下
	@classmethod
	def outputResult(cls, inputfile, outputfile):
		wb2 = openpyxl.Workbook()
		wb2.save('..\Data\\' + outputfile + '.xlsx')
		# print('新建成功')

		# 读取数据
		wb1 = openpyxl.load_workbook('..\Data\\' + inputfile + '.xlsx')
		wb2 = openpyxl.load_workbook('..\Data\\' + outputfile + '.xlsx')
		sheets1 = wb1.get_sheet_names()  # 获取sheet页
		sheets2 = wb2.get_sheet_names()
		sheet1 = wb1.get_sheet_by_name(sheets1[0])
		sheet2 = wb2.get_sheet_by_name(sheets2[0])

		max_row = sheet1.max_row  # 最大行数
		max_column = sheet1.max_column  # 最大列数

		for m in range(1, max_row + 1):
			for n in range(97, 97 + max_column):  # chr(97)='a'
				n = chr(n)  # ASCII字符
				i = '%s%d' % (n, m)  # 单元格编号
				cell1 = sheet1[i].value  # 获取data单元格数据
				sheet2[i].value = cell1  # 赋值到test单元格

		wb2.save('..\Data\\' + outputfile + '.xlsx')  # 保存数据

	# 获取row行column列的值
	@classmethod
	def getCell(cls, table, row, column):
		return table.cell(row, column).value

	# 获取表单
	@classmethod
	def getTable(cls, inputfile):
		data = xlrd.open_workbook('C:\Users\Administrator\PycharmProjects\WebTest\Data\\' + inputfile + '.xlsx')
		return data.sheets()[0]

	# float转datetime
	@classmethod
	def floatToDatetime(cls, cell):
		return datetime(*xldate_as_tuple(cell, 0))

	# 表单复制，添加result
	@classmethod
	def saveResult(cls, filename, row, column, flag, errorMsg):
		rb = xlrd.open_workbook('..\Data\\' + filename + '.xlsx')
		wb = copy(rb)
		ws = wb.get_sheet(0)
		if flag:
			ws.write(row, column, 'pass')
		else:
			ws.write(row, column, 'fail:' + errorMsg)
		wb.save('..\Data\\' + filename + '.xlsx')


# 网页操作类
class WebHandle():
	logger = Logger()

	@classmethod
	def __init__(cls, driver):
		cls.driver = driver

	# send_keys方法
	@classmethod
	def Input(cls, page, element, msg):
		el = cls.element(page, element)
		el.send_keys(msg)

	# click方法
	@classmethod
	def Click(cls, page, element):
		el = cls.element(page, element)
		el.click()

	# Zoe@2017-08-11
	# 一直等待某个元素消失，默认超时10秒
	def elementIsNotVisible(cls, page, element):
		try:
			el = WebDriverWait(UIHandle.driver, 10).until_not(
				EC.presence_of_element_located(local_config[page][element]))
			cls.logger.mylogger(page + element + 'is not display')
			return True
		except TimeoutException:
			cls.logger.mylogger(page + element + 'is displayed', -1)
			return False

	# element对象(还未完成。。。)
	def elements(cls, page, element):
		# 加入日志
		# cls.logger.loginfo(page)
		# 加入隐性等待
		WebDriverWait(cls.driver, 10)
		els = cls.driver.find_elements(*local_config[page][element])
		# 注意返回的是list
		return els

	# element对象（还可加入try，截图等。。。）
	@classmethod
	def element(cls, page, element):
		# 加入日志
		# cls.logger.mylogger(page)
		# 加入隐性等待
		# 此处便可以传入config_o1中的dict定位参数
		el = WebDriverWait(cls.driver, 10).until(EC.presence_of_element_located(local_config[page][element]))
		# 加入日志
		cls.logger.mylogger(page + element + 'OK')
		return el

	# 多种定位方式
	@classmethod
	def byId(cls, page, element):
		return cls.driver.find_element_by_id(local_config[page][element][1])

	@classmethod
	def byCssSelector(cls, page, element):
		return cls.driver.find_element_by_css_selector(local_config[page][element][1])

	@classmethod
	def byLinkText(cls, page, element):
		return cls.driver.find_element_by_link_text(local_config[page][element][1])

	@classmethod
	def byLinkTextByStr(cls, str):
		return cls.driver.find_element_by_link_text(str)

	@classmethod
	def byXPath(cls, page, element):
		return cls.driver.find_element_by_xpath(local_config[page][element][1])

	@classmethod
	def byTagName(cls, page, element):
		return cls.driver.find_element_by_tag_name(local_config[page][element][1])

	@classmethod
	def byIdAndXpath(cls, page, elementID, elementXpath):
		sub_element = cls.driver.find_element_by_id(local_config[page][elementID][1]).find_element_by_xpath(page,
																											elementXpath)
		ActionChains(cls).click(sub_element).perform()

	def byIdAndText(cls, page, elementID, elementText):
		sub_element = cls.driver.find_element_by_id(local_config[page][elementID][1]).find_element_by_link_text(page,
																												elementText)
		ActionChains(cls).click(sub_element).perform()

	# 定位下拉框
	@classmethod
	def SelectText(cls, v1):
		select = Select(cls.driver.find_element_by_tag_name('select'))
		select.select_by_visible_text(v1)

	# 另外一种方法选择下拉框,根据option的位置选择内容
	@classmethod
	def selectByIdNo(cls, page, element, num):
		cls.byId(page, element).find_elements_by_tag_name("option")[num].click()

	# 另外一种方法选择下拉框,根据option的位置选择内容

	@classmethod
	def selectByxpathNo(cls, page, element, num):
		cls.byXPath(page, element).find_elements_by_tag_name("option")[num].click()

	@classmethod
	def byXpathAndTagName(cls, page, element, tagname):
		sub_element = cls.byXPath(page, element).find_elements_by_tag_name(tagname)
		return sub_element

	# 分组数据选择
	# Kawy @2017-08-11
	# 修改传参，使函数可用于分层分析的内层和外层
	@classmethod
	def groupAddon(cls, webValue, excelValue, page, element1, element2):
		if excelValue < 0 or excelValue > 50:
			cls.logger.mylogger("case分组数据出错", -1)
		else:
			i = excelValue - webValue
			# 根据获取到的页面分组数值和列表中的数值求差值i来点击上或下
			while i != 0:
				if i > 0:
					cls.Click(page, element1)
					i -= 1
				else:
					cls.Click(page, element2)
					i += 1
				sleep(0.1)

	# 获取页面上的数据
	@classmethod
	def getText(cls, path):
		return cls.driver.find_element_by_xpath(path).text

	@classmethod
	def infoMatching(cls, filename, cellList, nrow, ncolumn, infoNum):
		flag = True
		errorMsg = ""
		for i in range(1, infoNum):
			if i >= 2:
				j = i + 1
			else:
				j = i
			flag = True
			str1 = str(cellList[i - 1])  # 表格信息
			str2 = str(cls.getText(
				"/html/body/div[2]/div[1]/div[3]/div/div/div[2]/div/div[2]/div[1]/div[1]/div/div[2]/ul/li[" + str(
					j) + "]/h2"))  # 网页信息
			if Assert.assertInfoEquals(str1, str2):  # 如果信息匹配
				continue  # 搜索下一个信息
			else:
				errorMsg = Assert.assertErrorType(i)
				cls.logger.mylogger(errorMsg, -1)
				flag = False
				break
		ExcelHandle.saveResult(filename, nrow, ncolumn, flag, errorMsg)

	# 返回element的text值
	@classmethod
	def getElementText(cls, page, element):
		return cls.byXPath(page, element).text

	# Kawy @2017-08-11
	# 选择分层分析的选择因子的函数
	@classmethod
	def selectFactorByXPath_SA(cls, uihandle, id, page, element):
		sub_element = uihandle.driver.find_element_by_id(id).find_element_by_xpath(local_config[page][element][1])
		ActionChains(uihandle.driver).click(sub_element).perform()

	# Kawy @2017-08-11
	# 选择分层分析的子因子的函数
	@classmethod
	def selectChildFactorByLinkText_SA(cls, uihandle, id, cell):
		sub2_element = uihandle.driver.find_element_by_id(id).find_element_by_link_text(cell)
		ActionChains(uihandle.driver).click(sub2_element).perform()

	# Kawy @2017-08-11
	# 选择分层分析的子因子的函数
	@classmethod
	def selectFactorByLinkText_SA(cls, uihandle, id, cell):
		sub2_element = uihandle.driver.find_element_by_id(id).find_element_by_xpath('//*[text()="' + cell + '"]')
		ActionChains(uihandle.driver).click(sub2_element).perform()

	# 股票池、基准、调仓频率、因子方向选择函数
	# Kawy @2017-08-14 Changed
	@classmethod
	def selectFactorInfo(cls, table, nrow, ncolumn, page, element):
		cell = ExcelHandle.getCell(table, nrow, ncolumn)
		cls.selectComboBox(page, element, str(cell))
		return cell

	# 点击股票池、基准、调仓频率、因子方向下拉框
	# Kawy @2017-08-14 Created
	@classmethod
	def selectComboBox(cls, page, element, factor):
		cls.byXPath(page, element).click()
		sleep(0.1)
		cls.driver.find_element_by_xpath(local_config[page][element][1] + "/*[text()='" + factor + "']").click()

	# 清除元素的内容
	# Zoe @2017-08-14
	@classmethod
	def clearText(cls, page, element):
		el = cls.element(page, element)
		el.clear()

	# 根据父元素2层定位子元素
	# Zoe @2017-08-14 Created
	@classmethod
	def loctorFromFather(cls, page, fa, sub, msg):
		el = cls.driver.find_element_by_xpath(local_config[page][fa][1]).find_element_by_xpath(
			local_config[page][sub][1])
		# 清空输入框的内容
		el.clear()
		# 输入msg
		el.send_keys(msg)

	#  鼠标悬停方法
	# Allen @2017-08-26 Created
	@classmethod
	def move_to_ele(cls, qqq, els):
		qqq = cls.driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[1]/div[1]/div[3]/div[1]/div/span")
		ActionChains(cls.driver).move_to_element(qqq).perform()
		els = cls.driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[1]/div[1]/div[3]/div[1]/div/div/i[1]')
		els.click()

	# 处理删除确认弹窗
	# Allen @2017-09-03 Creat
	@classmethod
	def alter1(cls):
		aa = cls.driver.switch_to_alter
		time.sleep(1)
		aa.accept()

	# Zoe @2017-08-28 Created
	'''根据产品管理中的table的xpath属性和table中的某一个元素定位其在table中的位置
	table从body开始
	queryContent：需要确定位置的内容
	column:表格的第几个元素，从0开始数
	page,element:table的xpath定位'''

	@classmethod
	def getTableContent(cls, queryContent, page, element, column, text):
		a = '\n'
		flag = False
		# 按行查询表格的数据，取出的数据是一整行，按空格分隔每一列的数据
		table_tr_list = cls.byXPath(page, element).text
		if a in table_tr_list:
			list = table_tr_list.split('\n')  # 以\n拆分成若干个(个数与列的个数相同)一维列表
			# 循环遍历table数据，确定查询数据的位置
			for i in range(len(list)):
				arr = list[i].split(" ")
				for j in range(len(arr)):
					if queryContent == arr[j]:
						WebHandle.compareTableContent(i + 1, column, text)
						Logger.mylogger(queryContent + '存在！')
						flag = True
		else:
			# 表格只有一行
			list = table_tr_list.split(" ")
			for i in range(len(list)):
				arr = list[i].split(" ")
				for j in range(len(arr)):
					if queryContent == arr[j]:
						WebHandle.compareTableContent(i + 1, column, text)
						Logger.mylogger(queryContent + '存在！')
						flag = True
		# 判读queryContent不存在，截图打印日志
		if not flag:
			Logger.mylogger(queryContent + '不存在！', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	# Zoe @2017-08-28 Created
	'''根据产品管理中的table的xpath属性和table中的某一个元素定位其在table中的位置并点击
	table从body开始
	queryContent：需要确定位置的内容
	 column:表格的第几个元素，从0开始数'''

	@classmethod
	def clickTableContent(cls, row, column):
		# if row==1:
		#   xpath = "//*[@id='tab1']/div[1]/table/tbody/tr/td[" + str(column) + "]/a"
		# else:
		if column == '1' or column == '2' or column == '3':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[" + str(column) + "]"
		if column == '5' or column == '7':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[" + str(column) + "]/a"
		# 出
		if column == '4-1':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[4]/a[1]"
		# 入
		if column == '4-2':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[4]/a[2]"
		# 启用
		if column == '6-1':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[6]/a[1]"
		# 禁用
		if column == '6-2':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[6]/a[2]"
		# 编辑
		if column == '8-1':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[8]/a[1]"
		# 删除
		if column == '8-2':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[8]/a[2]"
		# print arr
		cls.driver.find_element_by_xpath(xpath).click()

	# Zoe @2017-08-28 Created

	'''根据产品管理中的table的xpath属性和table中的某一个元素定位其在table中的位置，点击要跳转的元素
	table从body开始
	queryContent：需要确定位置的内容
	column:表格的第几个元素，从0开始数
	page,element:table的xpath定位'''

	@classmethod
	def swithTableContent(cls, queryContent, page, element, column):
		a = '\n'
		flag = False
		# 按行查询表格的数据，取出的数据是一整行，按空格分隔每一列的数据
		table_tr_list = cls.byXPath(page, element).text
		if a in table_tr_list:
			list = table_tr_list.split('\n')  # 以\n拆分成若干个(个数与列的个数相同)一维列表
			# 循环遍历table数据，确定查询数据的位置
			for i in range(len(list)):
				arr = list[i].split(" ")
				for j in range(len(arr)):
					if queryContent == arr[j]:
						# 根据需要查找的元素所在行位置点击要跳转的元素
						WebHandle.clickTableContent(i + 1, column)
						Logger.mylogger(queryContent + '存在！')
						flag = True
		else:
			# 表格只有一行
			list = table_tr_list.split(" ")
			for i in range(len(list)):
				arr = list[i].split(" ")
				for j in range(len(arr)):
					if queryContent == arr[j]:
						WebHandle.clickTableContent(i + 1, column)
						Logger.mylogger(queryContent + '存在！')
						flag = True
		# 判读queryContent不存在，截图打印日志
		if not flag:
			Logger.mylogger(queryContent + '不存在！', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

		# Zoe @2017-08-28 Created
		'''根据产品管理中的table的xpath属性和table中的某一个元素定位其在table中的位置，并检查其内容是否正确
		table从body开始
		queryContent：需要确定位置的内容
		column:表格的第几个元素，从0开始数
		text:需要比较的内容'''

	@classmethod
	def compareTableContent(cls, row, column, text):
		# if row == 1:
		#   xpath = "//*[@id='tab1']/div[1]/table/tbody/tr/td[" + str(column) + "]/a"
		# else:
		if column == 1 or column == 2 or column == 3:
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[" + str(column) + "]"
		if column == 5 or column == 7:
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[" + str(column) + "]/a"
		# 出
		if column == '4-1':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[4]/a[1]"
		# 入
		if column == '4-2':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[4]/a[2]"
		# 启用
		if column == '6-1':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[6]/a[1]"
		# 禁用
		if column == '6-2':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[6]/a[2]"
		# 编辑
		if column == '8-1':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[8]/a[1]"
		# 删除
		if column == '8-2':
			xpath = "//*[@id='tab1']/div[1]/table/tbody/tr[" + str(row) + "]/td[8]/a[2]"
		arr = cls.driver.find_element_by_xpath(xpath).text
		# print arr
		Assert.assertInfoEquals(arr, text)

	'''Zoe @2017-09-1 Created
	根据唯一的text以及xpath属性和table中的tr定位其在table中的位置，并选中此元素
	 page, element:从table的body定位
	queryContent：需要选中的位置的唯一内容
   '''

	@classmethod
	def selectTableByText(cls, queryContent, page, element):
		falg = False
		table_tr_list = []
		# 选中table下所有的tr开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name("tr")
		# for选好获取tr元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 根据需要查找的元素所在行位置点击要跳转的元素
			if queryContent in table_text:
				table_tr_list[i].click()
				Logger.mylogger(queryContent + '存在！')
				falg = True
				break
		if not falg:
			Logger.mylogger(queryContent + '不存在！', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	'''Zoe @2017-09-3 Created
	  根据xpath属性和table中的tr定位，并选中所有的元素
	   page, element:xpath选取表格
	 '''

	@classmethod
	def selectTableByAll(cls, page, element):
		table_tr_list = []
		# 选中table下所有的tr开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name("tr")
		# for选好获取tr元素的值
		for i in range(len(table_tr_list)):
			# table_text = table_tr_list[i].text
			# 点击要跳转的元素
			table_tr_list[i].click()

	'''Kawy @2017-09-04 Created
	根据唯一的text以及xpath属性和table中的tr定位其在table中的位置，并返回此元素
	 page, element:从table的body定位
	 row, column:要返回值的位置（行列）,从0开始算
	 maxcolumn:有多少列，从1开始算
   '''

	@classmethod
	def getTableElementByPosition(cls, page, element, row, column, maxcolumn):
		falg = False
		table_tr_list = []
		# 选中table下所有的tr开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).text
		# for选好获取tr元素的值
		position = row * maxcolumn + column
		list = table_tr_list.split('\n')  # 以\n拆分成若干个(个数与列的个数相同)一维列表
		# 循环遍历table数据，确定查询数据的位置
		arr = []
		for i in range(len(list)):
			arr += list[i].split(" ")
		if position <= len(arr):
			for i in range(len(arr)):
				if i == position:
					return arr[i]
		Logger.mylogger('元素不存在！', -1)
		html = Logger.creatHtml(picpath, pics)
		Logger.mylogger(html)
		raise Exception

	'''Zoe @2017-09-13 Created
	适用于报告管理中新建报告选择下拉框的内容
	根据唯一的text以及xpath属性定位位置，并选中此元素
	 page, element:从table的body定位
	queryContent：需要选中的位置的唯一内容
   '''

	@classmethod
	def selectReoprtByText(cls, queryContent, page, element):
		falg = False
		table_tr_list = []
		# 选中table下所有的tr开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name("div")
		# for选好获取div元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 根据需要查找的元素所在行位置点击要跳转的元素
			if queryContent in table_text:
				# 鼠标悬停
				ActionChains(cls.driver).move_to_element(table_tr_list[i]).perform()
				# 点击
				table_tr_list[i].click()
				Logger.mylogger(queryContent + '存在！')
				falg = True
				break
		if not falg:
			Logger.mylogger(queryContent + '不存在！', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	# 鼠标悬停方法
	# page, element是悬停元素的路径
	# Kawy @2017-09-12 Created
	@classmethod
	def moveToElement(cls, page, element):
		path = local_config[page][element][1]
		e = cls.driver.find_element_by_xpath(path)
		ActionChains(cls.driver).move_to_element(e).perform()

	# Kawy @2017-09-14
	# 鼠标右击方法，page和element定位到要右击的元素路径
	@classmethod
	def rightClick(cls, page, element):
		path = local_config[page][element][1]
		e = cls.driver.find_element_by_xpath(path)
		ActionChains(cls.driver).context_click(e).perform()

		'''Zoe @2017-09-14 Created
		 适用于策略管理列表中根据策略名称选择策略，点击策略中的详情，回测记录，重启，加入模拟等按钮
		  page, element:从table的body定位
		 queryContent：需要选中的位置的唯一内容
		 type:1-点击详情，2-回测记录，3-重启，4-加入模拟
		'''

	@classmethod
	def selectStrategyByText(cls, queryContent, page, element, type=1):
		falg = False
		table_tr_list = []
		# 选中table下所有的div开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name("div")
		# for选好获取div元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 根据需要查找的元素所在行位置点击要跳转的元素
			if queryContent in table_text:
				if type == 1:
					# 点击详情
					xpath = "//*[@id='wrapper']/div[3]/div[1]/div[" + str(i) + "]/div/div[1]/div"
					cls.driver.find_element_by_xpath(xpath).click()
				if type == 2:
					# 点击回测记录
					xpath = "//*[@id='wrapper']/div[3]/div[1]/div[" + str(i) + "]/div/div[4]/div[1]/div[1]"
					cls.driver.find_element_by_xpath(xpath).click()
				if type == 3:
					# 点击                    重启
					xpath = "//*[@id='wrapper']/div[3]/div[1]/div[" + str(i) + "]/div/div[4]/div[1]/div[2]"
					cls.driver.find_element_by_xpath(xpath).click()
				if type == 4:
					# 点击                    加入模拟
					xpath = "//*[@id='wrapper'/div[3]/div[1]/div[" + str(i) + "]/div/div[4]/div[1]/div[3]/span[1]/a"
					cls.driver.find_element_by_xpath(xpath).click()
				Logger.mylogger(queryContent + '存在！')
				falg = True
				break
		if not falg:
			Logger.mylogger(queryContent + '不存在！', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	'''Zoe @2017-09-14 Created
		适用于产品管理列表中根据策略名称选择产品，点击详情和业绩报告按钮
		 page, element:从table的body定位
		queryContent：需要选中的位置的唯一内容
		type:1-点击详情，2-业绩报告
	   '''

	@classmethod
	def selectProductByText(cls, queryContent, page, element, type=1):
		falg = False
		table_tr_list = []
		# 选中table下所有的div开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name("div")
		# for选好获取div元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 根据需要查找的元素所在行位置点击要跳转的元素
			if queryContent in table_text:
				if type == 1:
					# 点击详情
					xpath = "//*[@id='wrapper']/div[3]/div[1]/div[" + str(i) + "/div/div[4]/div/div[2]/a"
					cls.driver.find_element_by_xpath(xpath).click()
				if type == 2:
					# 点击回测记录
					xpath = "//*[@id='wrapper']/div[3]/div[1]/div[" + str(i) + "]/div/div[4]/div/div[1]/a"
					cls.driver.find_element_by_xpath(xpath).click()
				Logger.mylogger(queryContent + '存在！')
				falg = True
				break
		if not falg:
			Logger.mylogger(queryContent + '不存在！', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	# Zoe     @2017-09-14 Created 上传文件，path表示绝对路径
	@classmethod
	def uploadfile(cls, path):
		cls.driver.find_element_by_name('name').send_keys(path)

	'''Zoe @2017-09-14 Created
		适用于创建产品，根据行列填入数据
		 page, element:从table的body定位
		queryContent：其他账号，股票账号，期货账号，外汇账号
		type:1-点击详情，2-业绩报告
	   '''

	@classmethod
	def InputAccountByType(cls, queryContent, page, element, amount, dtime, accountname):
		falg = False
		table_tr_list = []
		# 选中table下所有的div开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name("div")
		# for选好获取div元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 根据需要查找的元素所在行位置点击要跳转的元素
			if queryContent in table_text:
				# 输入金额
				xpath = '//*[@id="wrapper"]/div[3]/div/div[2]/div/div/div[2]/div[' + str(i) + ']'
				cls.driver.find_element_by_xpath(xpath).find_element_by_id('amount').send_keys(amount)
				cls.driver.find_element_by_xpath(xpath).find_element_by_id('SubAccount').send_keys(dtime)
				cls.driver.find_element_by_xpath(xpath).find_element_by_id('SubAccount').send_keys(accountname)
				Logger.mylogger(queryContent + '存在！')
				falg = True
				break
		if not falg:
			Logger.mylogger(queryContent + '不存在！', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

		# Zoe @2017-08-28 Created

	'''账号费率设置根据行列填入费率的数据
	table从body开始
	row：第几行,从0开始
	column:表格的第几个元素，从1开始数
	page,element:table的xpath定位'''

	@classmethod
	def inputTableContent(cls, row, column, page, element, text):
		# 获取多少行
		row_list = cls.byXPath(page, element).find_elements_by_tag_name("tr")
		# 获取多少列
		col_list = cls.byXPath(page, element).find_elements_by_tag_name("td")
		# 如果行大于0，则获取每行有多少列数据
		if len(row_list) > 0:
			collist = len(col_list) / len(row_list)
		# 循环获取行数据
		for i in range(len(row_list)):
			if i == row:
				# 循环获取列数据
				for j in range(collist):
					if column == j:
						x = str(i + 1)
						# 如果只有一行path
						if len(row_list) == 1:
							xpath = '//*[@id="editTaxsModal3"]/div/div/div[2]/div[1]/div[2]/table/tbody/tr/td[' + str(
								j) + ']/div/input'
						else:
							# 多行的xpath
							xpath = '//*[@id="editTaxsModal3"]/div/div/div[2]/div[1]/div[2]/table/tbody/tr[' + x + ']/td[' + str(
								j) + ']/div/input'
						ele = cls.driver.find_element_by_xpath(xpath)
						ele.clear()
						ele.send_keys(text)
						break

	'''Kawy @2017-09-21 Created
	 根据唯一的text以及xpath属性和ul中的li定位其在ul中的位置，并选中此元素
	  page, element:从ul的body定位
	 queryContent：需要选中的位置的唯一内容
	'''

	@classmethod
	def selectULByText(cls, queryContent, page, element):
		falg = False
		table_tr_list = []
		# 选中ul下所有的li开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name("li")
		# for选好获取tr元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 根据需要查找的元素所在行位置点击要跳转的元素
			if queryContent in table_text:
				table_tr_list[i].click()
				Logger.mylogger(queryContent + '存在！')
				falg = True
				break
		if not falg:
			Logger.mylogger(queryContent + '不存在！', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	'''Kawy @2017-09-21 Created
		适用于创建组合，根据行列填入数据
		 page, element：从div的body定位
		queryContent：其他，股票，指数，期货，外汇
		account：初始本金
		date：入金时间
	   '''

	@classmethod
	def InputAccountByDiv(cls, queryContent, page, element, account):
		falg = False
		table_tr_list = []
		# 选中table下所有的div开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name("span")
		# for选好获取div元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 根据需要查找的元素所在行位置点击要跳转的元素
			if queryContent in table_text:
				# 输入金额
				# amount=cls.driver.find_element_by_xpath('//*[@id="newProduct"]/div[1]/div[2]/div/div['+str(i+1)+']/div').find_element_by_id('InitAmount')
				cls.driver.find_element_by_xpath(
					'//*[@id="newProduct"]/div[1]/div[2]/div/div[' + str(i + 1) + ']/div').find_element_by_id(
					'InitAmount').clear()
				cls.driver.find_element_by_xpath(
					'//*[@id="newProduct"]/div[1]/div[2]/div/div[' + str(i + 1) + ']/div').find_element_by_id(
					'InitAmount').send_keys(account)
				Logger.mylogger(queryContent + '存在！')
				falg = True
				break
		if not falg:
			Logger.mylogger(queryContent + '不存在')

	'''Kawy @2017-09-25 Created
		适用于Alpha策略-信号生成，根据名称点击加入信号
		 page, element：从div的body定位
		queryContent：信号生成策略名称
	   '''

	@classmethod
	def clickAddSignalByDiv(cls, queryContent, page, element):
		flag = False
		table_tr_list = []
		# 选中table下所有的div开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name("span")
		# for选好获取div元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 根据需要查找的元素所在行位置点击要跳转的元素
			if queryContent in table_text:
				# 输入金额
				path = '//*[@id="alpha_list' + str(i) + '"]/div[1]/div[2]/div[2]/a[3]'
				cls.driver.find_element_by_xpath(path).click()
				Logger.mylogger(queryContent + '存在！')
				flag = True
				break
		if not flag:
			Logger.mylogger(queryContent + '不存在', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception

	# 鼠标悬停方法，适用于因子排名，点击分析按钮
	# page, element是悬停元素的路径
	# Zoe @2017-09-27 Created
	@classmethod
	def moveToElementByJS(cls, page, element):
		path = local_config[page][element][1]
		# 选取第一个tag 是rect的
		btn = cls.driver.find_element_by_xpath(path).find_element_by_tag_name("rect")
		sleep(1)
		# 调用js鼠标悬停时间
		cls.driver.execute_script("$(arguments[0]).mouseenter()", btn)
		sleep(2)

	# kawy @2017-10-25 Created
	# 查找tagname并点击
	@classmethod
	def selectTagNameByText(cls, queryContent, page, element, tagname='li'):
		falg = False
		table_tr_list = []
		# 选中table下所有的tagname开头的元素赋值到列表中
		table_tr_list = cls.byXPath(page, element).find_elements_by_tag_name(tagname)
		# for选好获取tagname元素的值
		for i in range(len(table_tr_list)):
			table_text = table_tr_list[i].text
			# 根据需要查找的元素所在行位置点击要跳转的元素
			if queryContent in table_text:
				table_tr_list[i].click()
				Logger.mylogger(queryContent + '存在！')
				falg = True
				break
		if not falg:
			Logger.mylogger(queryContent + '不存在！', -1)
			html = Logger.creatHtml(picpath, pics)
			Logger.mylogger(html)
			raise Exception


# 用户界面操作类
class UIHandle():
	logger = Logger()

	# 构造方法，用来接收selenium的driver对象
	@classmethod
	def __init__(cls, driver):
		cls.driver = driver

	# 输入地址
	@classmethod
	def get(cls, url):
		# cls.logger.loginfo(url)
		cls.driver.get(url)

	# 关闭浏览器驱动
	@classmethod
	def quit(cls):
		cls.driver.quit()

	# 截图
	@classmethod
	def getScreenshots(cls, filename):
		cls.scrollUpScreen()
		cls.get_screenshot_as_file('..\Snapshot\\' + filename + '(1).jpg')
		cls.scrollDownScreen()
		cls.get_screenshot_as_file('..\Snapshot\\' + filename + '(2).jpg')
		cls.scrollUpScreen()

	@classmethod
	def get_screenshot_as_file(cls, filename):
		cls.driver.get_screenshot_as_file(filename)

	# 下拉
	@classmethod
	def scrollDownScreen(cls):
		cls.driver.execute_script("""
       (function () {
        var y = 0;
        var step = 100;
        var dom=document.getElementsByClassName("content")[0];
        dom.scrollTop=0;
        function f() {
            if (y < dom.scrollHeight) {
                y += step;
                dom.scrollTop=y;
                setTimeout(f, 100);
            }else{
            document.title += "scroll-done";
            }

        }
          f();
       })();
       """)
		sleep(2)

	# 上拉
	@classmethod
	def scrollUpScreen(cls):
		cls.driver.execute_script("""
       (function () {
        var dom=document.getElementsByClassName("content")[0];
        var y = dom.scrollHeight;
        var step = 100;
        dom.scrollTop=dom.scrollHeight;
        function f() {
            if (y > 0) {
                y -= step;
                dom.scrollTop=y;
                setTimeout(f, 100);
            }else{
            document.title += "scroll-done";
            }

        }
          f();
       })();
       """)
		sleep(2)

	# Zoe,created@2017-09-03
	# 关闭浏览器子窗口驱动
	@classmethod
	def quitNewWindow(cls):
		# 获取当前所有窗口句柄（窗口A、B）
		handles = cls.driver.window_handles
		# 对窗口进行遍历
		for newhandle in handles:
			# 关闭新打开的窗口
			if newhandle != handles[0]:
				cls.driver.switch_to_window(newhandle)
				cls.driver.close()
		# 返回主窗口
		cls.driver.switch_to_window(handles[0])

